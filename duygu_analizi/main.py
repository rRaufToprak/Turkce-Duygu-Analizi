# Türkçe duygu analizi
# Rauf Toprak

import numpy as np
import pandas as pd
import itertools
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import PassiveAggressiveClassifier
from sklearn.metrics import accuracy_score, confusion_matrix
from openpyxl import Workbook,load_workbook
from TurkishStemmer import TurkishStemmer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook("comment.xlsx")
ws = wb.active
df = ws


wb = Workbook()
ws = wb.active


df = pd.DataFrame()
wb = load_workbook(filename = 'comment.xlsx')
sheet_names = wb.get_sheet_names()
name = sheet_names[0]
sheet_ranges = wb[name]
df = pd.DataFrame(sheet_ranges.values)
data = sheet_names

df = df.rename(columns={0: 'site', 1: 'product', 2: 'comment', 3: 'point', 4: 'label'})

df.drop(df[df["label"] == 3].index, inplace=True)
df["label"] = df["label"].replace(1, 'negative')
df["label"] = df["label"].replace(2, 'negative')
df["label"] = df["label"].replace(4, 'positive')
df["label"] = df["label"].replace(5, 'positive')
df["label"].value_counts()

df["comment"] = df["comment"].apply(lambda x: x.lower())


def remove_numeric(corpus):
    output = "".join(words for words in corpus if not words.isdigit())
    return output


df["comment"] = df["comment"].apply(lambda x: remove_numeric(x))


def remove_stopwords(df_fon):
    stopwords = open('turkce-stop-words', 'r').read().split()
    df_fon['stopwords_removed'] = list(map(lambda doc:
                                           [word for word in doc if word not in stopwords], df_fon['comment']))


remove_stopwords(df)

stemmer = TurkishStemmer()

df['comment'] = df['comment'].apply(lambda x: stemmer.stem(x))

df = df.replace([np.inf, -np.inf], np.nan)
df = df.dropna()
df = df.reset_index()

labels = df.label
x_train, x_test, y_train, y_test = train_test_split(df['comment'], labels, test_size=0.2, random_state=7)
tfidf_vectorizer = TfidfVectorizer(max_df=0.7)

tfidf_train = tfidf_vectorizer.fit_transform(x_train)
tfidf_test = tfidf_vectorizer.transform(x_test)
pac = PassiveAggressiveClassifier(max_iter=50)
pac.fit(tfidf_train, y_train)
y_pred = pac.predict(tfidf_test)
score = accuracy_score(y_test, y_pred)
print(f'Accuracy: {round(score * 100, 2)}%')
confusion_matrix(y_test, y_pred, labels=['positive', 'negative'])
print(tfidf_train.toarray())


def duygu_analizi():
    yorum = input("Yorum giriniz: ")
    tfidf_test = tfidf_vectorizer.transform([yorum])

    y_pred = pac.predict(tfidf_test.toarray())
    print(y_pred)


while True:
    duygu_analizi()
